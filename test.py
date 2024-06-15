from openpyxl import Workbook
import smtpd


wb = Workbook()
ws = wb.active

# cabeçalho
ws['A1'] = 'Tipos de Ataque'
ws['B1'] = 'Descrição'
ws['C1'] = 'ID'
ws['D1'] = 'Plataforma executada'
ws['E1'] = 'Codígo CVE'
ws['F1'] = 'IP do servidor'
ws['G1'] = 'Protocolo/Porta'
ws['H1'] = 'Versão do Serviço'
ws['I1'] = 'Arquivos infectados'
ws['J1'] = 'Níveis de urgência'



# conteudo ...
ws['A2'] = ''



wb.save('exemplo.xlsx')


import smtplib
from email.mime.text import MIMEText

# Configurações do servidor SMTP
smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_user = 'seu-email@gmail.com'
smtp_password = 'sua-senha'

# Endereço de e-mail do Zapier
zapier_email = 'your-zap-email@zapiermail.com'

# Dados do e-mail
subject = 'Novo dado do Python'
body = 'Aqui estão os dados enviados pelo script Python.'
msg = MIMEText(body)
msg['Subject'] = subject
msg['From'] = smtp_user
msg['To'] = zapier_email

# Enviando o e-mail
try:
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(smtp_user, smtp_password)
    server.sendmail(smtp_user, [zapier_email], msg.as_string())
    server.quit()
    print('E-mail enviado com sucesso!')
except Exception as e:
    print(f'Erro ao enviar e-mail: {e}')




